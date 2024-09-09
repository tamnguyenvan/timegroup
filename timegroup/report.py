from openpyxl import Workbook
from datetime import datetime


def format_updated_at(updated_at_str):
    dt = datetime.fromisoformat(updated_at_str)
    return dt.strftime("%d/%m/%Y %H:%M")

class Report:
    def __init__(self):
        self._data = {}

    @property
    def data(self):
        return self._data

    @staticmethod
    def merge(self, *reports):
        for report in reports:
            self._data.update(report.data)
        return self

    def parse(self, raw_data):
        raise NotImplementedError

    def save(self, outfile):
        wb = Workbook()

        for sheet_name, sheet_data in self._data.items():
            # Load the existing workbook
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]  # Use existing sheet
            else:
                ws = wb.create_sheet(title=sheet_name)

            if len(sheet_data) == 0:
                continue

            for row in sheet_data:
                ws.append(row)

        wb.save(outfile)

class POSReport(Report):
    def __init__(self):
        super().__init__()
        self._column_names = ["Ngày tạo đơn", "COD", "Tổng số lượng SP", "Số lượng", "Mã sản phẩm", "Mã Mẫu Mã"]
        self._active = "POS"
        self._data = {"POS": [self._column_names]}

    def parse(self, raw_data):
        data = self._data[self._active]

        for order in raw_data:
            items = order.get("items", [])
            for i, item in enumerate(items):
                row = ["" for _ in range(len(self._column_names))]
                if i == 0:
                    # Created date
                    status_history = order.get("status_history", [])
                    created_order_date = None
                    for status_data in status_history:
                        if status_data.get("status") == 1:
                            created_order_date = status_data["updated_at"]

                    if created_order_date is None:
                        continue

                    row[0] = format_updated_at(created_order_date)
                    row[1] = order["cod"]
                    row[2] = order["total_quantity"]

                row[3] = item["quantity"]
                variation_info = item["variation_info"]
                row[4] = variation_info["product_display_id"]
                # name = variation_info["name"]
                # if variation_info["fields"]:
                #     row[5] = " / ".join([name] + [field["name"] + ": " + field["value"] for field in variation_info["fields"]])
                row[5] = variation_info["display_id"]

                data.append(row)
        return len(data)

class RemainProductReport(Report):
    def __init__(self):
        super().__init__()
        self._column_names = [
            "MA_SP", "MA_MAU_MA", "TON_KHO", "Danh mục", "Tổng nhập"
        ]
        self._active = "TỒN KHO"
        self._data = {"TỒN KHO": [self._column_names]}

    def parse(self, raw_data):
        data = self._data[self._active]
        for variation in raw_data:
            row = ["" for _ in range(len(self._column_names))]

            product_info = variation.get("product")
            if product_info:
                # ma_sp
                row[0] = product_info["display_id"]

                # danh_muc
                categories = product_info.get("categories")
                if categories:
                    row[3] = categories[0].get("name", "")

            # ma_mau_ma
            row[1] = variation["display_id"]

            variations_warehouses = variation.get("variations_warehouses", [])
            if variations_warehouses:
                warehouse = variations_warehouses[0]
                row[2] = warehouse["actual_remain_quantity"]
                row[4] = warehouse["total_quantity"]

            data.append(row)
        return len(data)


class AwaitingOrderReport(Report):
    def __init__(self):
        super().__init__()
        self._column_names = [
            "Ngày tạo đơn", "Mã sản phẩm", "Mã Mẫu mã", "Sản phẩm", "Mã đơn hàng",
            "Tổng số lượng SP", "Số lượng", "Giá", "Giảm giá", "Tình trạng kho", "COD"
        ]
        self._active = "CHỜ HÀNG"
        self._data = {"CHỜ HÀNG": [self._column_names]}

    def parse(self, raw_data):
        data = self._data[self._active]
        for order in raw_data:
            items = order.get("items", [])
            for i, item in enumerate(items):
                row = ["" for _ in range(len(self._column_names))]
                if i == 0:
                    # Created date
                    status_history = order.get("status_history", [])
                    created_order_date = None
                    for status_data in status_history:
                        if status_data.get("status") == 1:
                            created_order_date = status_data["updated_at"]

                    if created_order_date is None:
                        continue

                    # ngay_tao_don
                    row[0] = format_updated_at(created_order_date)

                    # ma_don_hang
                    row[4] = order["id"]

                    # tong_so_luong
                    row[5] = order["total_quantity"]

                    # cod
                    row[10] = order["cod"]

                variation_info = item["variation_info"]
                # ma_san_pham
                row[1] = variation_info["product_display_id"]

                # ma_mau_ma
                row[2] = variation_info["display_id"]

                # san_pham
                name = variation_info["name"]
                if variation_info["fields"]:
                    row[3] = " / ".join([name] + [field["name"] + ": " + field["value"] for field in variation_info["fields"]])

                # so_luong
                row[6] = item["quantity"]

                # gia
                row[7] = variation_info["retail_price"]

                # giam gia
                row[8] = item["discount_each_product"]

                row[9] = ""
                data.append(row)

        return len(data)


class GHTKOrderReport(Report):
    def __init__(self):
        super().__init__()
        self._partner_id = 1
        self._partner_name = "GHTK"
        self._column_names = [
            "MVĐ", "Khách hàng", "SĐT", "Mã sản phẩm", "Mã mẫu mã", "Sản phẩm", "Tổng SL",
            "Số lượng", "COD", "Phí trả cho ĐVVC", "Facebook Page", "PAGE ID",
            "Người xử lý", "NV xác nhận", "Ngày tạo đơn", "Kho hàng", "Ngày gửi"
        ]
        self._active = "Đơn Hàng GHTK"
        self._data = {"Đơn Hàng GHTK": [self._column_names]}

    def parse(self, raw_data):
        data = self._data[self._active]
        partner_ids = set()
        for order in raw_data:
            partner = order.get("partner") or {}
            partner_id = partner.get("partner_id")
            partner_ids.add(partner_id)
            if partner_id != self._partner_id:
                continue

            items = order.get("items", [])
            for i, item in enumerate(items):
                row = ["" for _ in range(len(self._column_names))]
                if i == 0:
                    # ma van don
                    extend_code = partner.get("extend_code", "")

                    # don vi van chuyen
                    row[0] = extend_code

                    # Created date
                    status_history = order.get("status_history", [])
                    created_order_date = None
                    for status_data in status_history:
                        if status_data.get("status") == 1:
                            created_order_date = status_data["updated_at"]

                    if created_order_date is None:
                        continue

                    row[0] = extend_code
                    row[14] = format_updated_at(created_order_date)
                    row[8] = order["cod"]
                    row[6] = order["total_quantity"]

                    # khach hang
                    customer = order["customer"]
                    customer_name = customer["name"]
                    row[1] = customer_name

                    # sdt
                    phone_number = (customer.get("phone_numbers") or [""])[0]
                    row[2] = phone_number

                    # phi tra cho DVVC
                    row[9] = order.get("partner_fee", "")

                    page = order["page"]
                    facebook_page = page["name"]
                    page_id = page["id"]
                    row[10] = facebook_page
                    row[11] = page_id

                    # nguoi xu ly
                    assigning_seller = order["assigning_seller"]
                    row[12] = assigning_seller["name"]

                    # nguoi xac nhan
                    confirm_staff = ""
                    status_history = order.get("status_history", [])
                    for status_item in status_history:
                        if status_item["status"] in (1, 11):
                            confirm_staff = status_item["name"]
                    row[13] = confirm_staff

                    # kho hang
                    warehouse_info = order["warehouse_info"]
                    row[15] = warehouse_info["name"]

                # ma_san_pham
                variation_info = item["variation_info"]
                row[3] = variation_info["product_display_id"]

                # ma_mau_ma
                row[4] = variation_info["display_id"]

                # san_pham
                name = variation_info["name"]
                if variation_info["fields"]:
                    row[5] = " / ".join([name] + [field["name"] + ": " + field["value"] for field in variation_info["fields"]])

                # so luong
                row[7] = item["quantity"]

                data.append(row)
        return len(data)

class VTPOrderReport(Report):
    def __init__(self):
        super().__init__()
        self._partner_id = 3
        self._column_names = [
            "MVĐ", "Khách hàng", "SĐT", "Mã sản phẩm", "Mã mẫu mã", "Sản phẩm", "Tổng SL",
            "Số lượng", "COD", "Phí trả cho ĐVVC", "Facebook Page", "PAGE ID",
            "Người xử lý", "NV xác nhận", "Ngày tạo đơn", "Kho hàng", "Ngày gửi"
        ]
        self._active = "Đơn Hàng VTP"
        self._data = {"Đơn Hàng VTP": [self._column_names]}

    def parse(self, raw_data):
        data = self._data[self._active]
        for order in raw_data:
            partner = order.get("partner")

            # ma van don
            extend_code = partner.get("extend_code")

            # don vi van chuyen
            partner_id = partner.get("partner_id")

            items = order.get("items", [])
            for i, item in enumerate(items):
                row = ["" for _ in range(len(self._column_names))]
                if i == 0:
                    # Created date
                    status_history = order.get("status_history", [])
                    created_order_date = None
                    for status_data in status_history:
                        if status_data.get("status") == 1:
                            created_order_date = status_data["updated_at"]

                    if created_order_date is None:
                        continue

                    row[0] = extend_code
                    row[14] = format_updated_at(created_order_date)
                    row[8] = order["cod"]
                    row[6] = order["total_quantity"]

                    # khach hang
                    customer = order["customer"]
                    customer_name = customer["name"]
                    row[1] = customer_name

                    # sdt
                    phone_number = (customer.get("phone_numbers") or [""])[0]
                    row[2] = phone_number

                # ma_san_pham
                variation_info = item["variation_info"]
                row[3] = variation_info["product_display_id"]

                # ma_mau_ma
                row[4] = variation_info["display_id"]

                # san_pham
                name = variation_info["name"]
                if variation_info["fields"]:
                    row[5] = " / ".join([name] + [field["name"] + ": " + field["value"] for field in variation_info["fields"]])

                # so luong
                row[7] = item["quantity"]

                # phi tra cho DVVC
                row[9] = item.get("partner_fee", "")

                page = order["page"]
                facebook_page = page["name"]
                page_id = page["id"]
                row[10] = facebook_page
                row[11] = page_id

                # nguoi xu ly
                assigning_seller = order["assigning_seller"]
                row[12] = assigning_seller["name"]

                # nguoi xac nhan
                confirm_staff = ""
                status_history = order.get("status_history", [])
                for status_item in status_history:
                    if status_item["status"] in (1, 11):
                        confirm_staff = status_item["name"]
                row[13] = confirm_staff

                # kho hang
                warehouse_info = order["warehouse_info"]
                row[15] = warehouse_info["name"]

                data.append(row)
        return len(data)